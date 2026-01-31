import csv
import json
from pathlib import Path
from typing import Dict, List, Tuple
from datetime import datetime

class Exporter:
    
    def __init__(self):
        self.last_export_path: Path = None
        self.last_export_paths: List[Path] = []
    
    def export_to_txt(self, abbreviations: Dict[str, dict], output_path: str) -> bool:
        
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_file, 'w', encoding='utf-8') as f:
                # Header
                f.write("=" * 70 + "\n")
                f.write("SCOUT ABBREVIATION REPORT\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 70 + "\n\n")
                
                # Statistics
                total = len(abbreviations)
                with_def = sum(1 for info in abbreviations.values() if info['definition'])
                
                f.write(f"Total Abbreviations: {total}\n")
                f.write(f"With Definitions: {with_def}\n")
                f.write(f"Without Definitions: {total - with_def}\n")
                f.write("\n" + "-" * 70 + "\n\n")
                
                # Abbreviations (sorted alphabetically)
                for abbrev in sorted(abbreviations.keys()):
                    info = abbreviations[abbrev]
                    definition = info['definition'] or "Definition not found"
                    count = info['count']
                    file_count = len(info['files'])
                    
                    # Format abbreviation entry
                    f.write(f"{abbrev}\n")
                    f.write(f"  Definition: {definition}\n")
                    f.write(f"  Occurrences: {count} time(s)\n")
                    f.write(f"  Found in: {file_count} file(s)\n")
                    
                    if info['files'] and file_count <= 3:
                        # Show all files if 3 or fewer
                        f.write(f"  Files: {', '.join(Path(fp).name for fp in info['files'])}\n")
                    elif info['files']:
                        # Show first 3 files with count of remaining
                        shown = ', '.join(Path(fp).name for fp in info['files'][:3])
                        remaining = file_count - 3
                        f.write(f"  Files: {shown} (and {remaining} more)\n")
                    
                    f.write("\n")
            
            self.last_export_path = output_file
            return True
            
        except Exception as e:
            print(f"Error exporting to TXT: {e}")
            return False
    
    def export_to_csv(self, abbreviations: Dict[str, dict], output_path: str) -> bool:
       
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow(['Abbreviation', 'Definition', 'Occurrences', 
                               'File Count', 'Files'])
                
                # Data rows (sorted alphabetically)
                for abbrev in sorted(abbreviations.keys()):
                    info = abbreviations[abbrev]
                    definition = info['definition'] or "Definition not found"
                    count = info['count']
                    file_count = len(info['files'])
                    files = '; '.join(Path(f).name for f in info['files'])
                    
                    writer.writerow([abbrev, definition, count, file_count, files])
            
            self.last_export_path = output_file
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def export_to_json(self, abbreviations: Dict[str, dict], output_path: str) -> bool:
        
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Prepare data structure
            export_data = {
                'metadata': {
                    'generated': datetime.now().isoformat(),
                    'total_abbreviations': len(abbreviations),
                    'with_definitions': sum(1 for info in abbreviations.values() 
                                          if info['definition']),
                },
                'abbreviations': {}
            }
            
            # Add abbreviations
            for abbrev, info in abbreviations.items():
                export_data['abbreviations'][abbrev] = {
                    'definition': info['definition'],
                    'occurrences': info['count'],
                    'files': [str(Path(f).name) for f in info['files']]
                }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            self.last_export_path = output_file
            return True
            
        except Exception as e:
            print(f"Error exporting to JSON: {e}")
            return False
    
    def export_to_excel(self, abbreviations: Dict[str, dict], output_path: str) -> bool:
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Abbreviations"
            
            # Header row with styling
            headers = ['Abbreviation', 'Definition', 'Occurrences', 'File Count', 'Files']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row, (abbrev, info) in enumerate(sorted(abbreviations.items()), 2):
                definition = info['definition'] or "Definition not found"
                count = info['count']
                file_count = len(info['files'])
                files = ', '.join(Path(f).name for f in info['files'])
                
                ws.cell(row=row, column=1, value=abbrev)
                ws.cell(row=row, column=2, value=definition)
                ws.cell(row=row, column=3, value=count)
                ws.cell(row=row, column=4, value=file_count)
                ws.cell(row=row, column=5, value=files)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 40
            
            wb.save(output_file)
            self.last_export_path = output_file
            return True
            
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_to_pdf(self, abbreviations: Dict[str, dict], output_path: str) -> bool:
        
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            doc = SimpleDocTemplate(str(output_file), pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
            )
            elements.append(Paragraph("Scout Abbreviation Report", title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Statistics
            total = len(abbreviations)
            with_def = sum(1 for info in abbreviations.values() if info['definition'])
            
            stats_style = ParagraphStyle(
                'Stats',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=12,
            )
            elements.append(Paragraph(f"<b>Total Abbreviations:</b> {total}", stats_style))
            elements.append(Paragraph(f"<b>With Definitions:</b> {with_def}", stats_style))
            elements.append(Paragraph(f"<b>Without Definitions:</b> {total - with_def}", stats_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Table data
            data = [['Abbreviation', 'Definition', 'Count', 'Files']]
            for abbrev, info in sorted(abbreviations.items()):
                definition = info['definition'] or "Definition not found"
                count = str(info['count'])
                file_count = str(len(info['files']))
                data.append([abbrev, definition[:80], count, file_count])
            
            # Create table
            table = Table(data, colWidths=[1.2*inch, 4*inch, 0.8*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            self.last_export_path = output_file
            return True
            
        except ImportError:
            print("reportlab not installed. Install with: pip install reportlab")
            return False
        except Exception as e:
            print(f"Error exporting to PDF: {e}")
            return False
    
    def export(self, abbreviations: Dict[str, dict], output_path: str, 
               format: str = 'txt') -> bool:
        
        format = format.lower()
        
        if format == 'txt':
            return self.export_to_txt(abbreviations, output_path)
        elif format == 'csv':
            return self.export_to_csv(abbreviations, output_path)
        elif format == 'json':
            return self.export_to_json(abbreviations, output_path)
        elif format in ['xlsx', 'excel']:
            return self.export_to_excel(abbreviations, output_path)
        elif format == 'pdf':
            return self.export_to_pdf(abbreviations, output_path)
        else:
            print(f"Unsupported format: {format}")
            return False
    
    def get_default_filename(self, format: str = 'txt') -> str:
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"scout_report_{timestamp}.{format}"
    
    def _split_abbreviations(self, abbreviations: Dict[str, dict], 
                            max_per_file: int) -> List[Dict[str, dict]]:
        """Split abbreviations into multiple batches"""
        if max_per_file <= 0 or len(abbreviations) <= max_per_file:
            return [abbreviations]
        
        batches = []
        items = list(sorted(abbreviations.items()))
        
        for i in range(0, len(items), max_per_file):
            batch = dict(items[i:i + max_per_file])
            batches.append(batch)
        
        return batches
    
    def export_batch(self, abbreviations: Dict[str, dict], output_path: str,
                    format: str = 'txt', limit: int = None, 
                    items_per_file: int = None) -> Tuple[bool, List[str]]:
        """
        Export abbreviations with options for limiting and splitting.
        
        Args:
            abbreviations: Dictionary of abbreviations to export
            output_path: Base path for output file(s)
            format: Export format (txt, csv, json, xlsx, pdf)
            limit: Maximum total abbreviations to export (None = all)
            items_per_file: Number of abbreviations per file (None = all in one file)
        
        Returns:
            tuple: (success: bool, created_files: List[str])
        """
        try:
            # Apply limit if specified
            if limit and limit > 0:
                sorted_items = sorted(abbreviations.items())[:limit]
                abbreviations = dict(sorted_items)
            
            # Split into batches if needed
            if items_per_file and items_per_file > 0:
                batches = self._split_abbreviations(abbreviations, items_per_file)
            else:
                batches = [abbreviations]
            
            # Reset export paths list
            self.last_export_paths = []
            created_files = []
            
            # Export each batch
            output_file = Path(output_path)
            base_name = output_file.stem
            extension = output_file.suffix or f'.{format}'
            parent_dir = output_file.parent
            
            for i, batch in enumerate(batches):
                if len(batches) > 1:
                    # Create numbered filename for multiple files
                    batch_filename = f"{base_name}_part{i+1}{extension}"
                    batch_path = parent_dir / batch_filename
                else:
                    batch_path = output_file
                
                # Export this batch
                success = self._export_single(batch, str(batch_path), format)
                
                if not success:
                    return False, created_files
                
                self.last_export_paths.append(batch_path)
                created_files.append(str(batch_path))
            
            # Set last export path to first file for compatibility
            if self.last_export_paths:
                self.last_export_path = self.last_export_paths[0]
            
            return True, created_files
            
        except Exception as e:
            print(f"Error in batch export: {e}")
            return False, []
    
    def _export_single(self, abbreviations: Dict[str, dict], 
                      output_path: str, format: str) -> bool:
        """Export a single batch of abbreviations"""
        format = format.lower()
        
        if format == 'txt':
            return self.export_to_txt(abbreviations, output_path)
        elif format == 'csv':
            return self.export_to_csv(abbreviations, output_path)
        elif format == 'json':
            return self.export_to_json(abbreviations, output_path)
        elif format in ['xlsx', 'excel']:
            return self.export_to_excel(abbreviations, output_path)
        elif format == 'pdf':
            return self.export_to_pdf(abbreviations, output_path)
        else:
            print(f"Unsupported format: {format}")
            return False
